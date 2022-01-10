import cv2
import pandas as pd
from openpyxl import load_workbook


def gstreamer_pipeline(
    capture_width=1280,
    capture_height=720,
    display_width=1280,
    display_height=720,
    framerate=30,
    flip_method=0,
):
    return (
        "nvarguscamerasrc ! "
        "video/x-raw(memory:NVMM), "
        "width=(int)%d, height=(int)%d, "
        "format=(string)NV12, framerate=(fraction)%d/1 ! "
        "nvvidconv flip-method=%d ! "
        "video/x-raw, width=(int)%d, height=(int)%d, format=(string)BGRx ! "
        "videoconvert ! "
        "video/x-raw, format=(string)BGR ! appsink"
        % (
            capture_width,
            capture_height,
            framerate,
            flip_method,
            display_width,
            display_height,
        )
    )

def rescale_frame(frame, percent):
    width = int(frame.shape[1] * percent/ 100)
    height = int(frame.shape[0] * percent/ 100)
    dim = (width, height)
    return cv2.resize(frame, dim, interpolation =cv2.INTER_AREA)

#cap = cv2.VideoCapture(1, cv2.CAP_V4L2)

h = s = v = xpos = ypos = 0

def draw_function(event, x, y, flags, frame):
    if event == cv2.EVENT_LBUTTONUP:
        global b, g, r, xpos, ypos, clicked
        # clicked = True
        xpos = x
        ypos = y
        imgHSV = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        h, s, v = imgHSV[y, x]
        print(h,s,v)
        data_hsv = {'h':[h],
                    's':[s],
                    'v':[v]}
        
        df_hsv = pd.DataFrame(data_hsv)
        print(df_hsv)
        #df_hsv.to_excel('hsv_ELP.xlsx', index=False) #create excel
        path = "hsv_ELP.xlsx"
        book = load_workbook(path)
        writer = pd.ExcelWriter("hsv_ELP.xlsx", engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df_hsv.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
        print("saving as hsv...")
        writer.save()
        
cv2.namedWindow('image')

while(True):
    print(gstreamer_pipeline(flip_method=0))
    cap = cv2.VideoCapture(gstreamer_pipeline(flip_method=0), cv2.CAP_GSTREAMER)
    # imgBlur = cv2.blur(ori, (5,5))
    if cap.isOpened():
        window_handle = cv2.namedWindow("image", cv2.WINDOW_AUTOSIZE)
        # Window
        while cv2.getWindowProperty("image", 0) >= 0:
            ret_val, img = cap.read()
            imgscale = rescale_frame(img, percent=40)
            cv2.imshow("image", imgscale)
            cv2.setMouseCallback('image', draw_function, imgscale)
            # This also acts as
            keyCode = cv2.waitKey(1) & 0xFF == ord('q')
            # Stop the program on the ESC key
            if keyCode == 27:
                break
        cap.release()
        cv2.destroyAllWindows()
    else:
        print("Unable to open camera")
    
    #cv2.imshow('image', imgscale)
    cv2.waitKey(1) 
    
cap.release()
cv2.destroyAllWindows()