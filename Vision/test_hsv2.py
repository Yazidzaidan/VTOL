import cv2
import pandas as pd
from openpyxl import load_workbook
cap = cv2.VideoCapture(1)

def rescale_frame(frame, percent):
    width = int(frame.shape[1] * percent/ 100)
    height = int(frame.shape[0] * percent/ 100)
    dim = (width, height)
    return cv2.resize(frame, dim, interpolation =cv2.INTER_AREA)

h = s = v = xpos = ypos = 0

def draw_function(event, x, y, flags, frame):
    if event == cv2.EVENT_LBUTTONUP:
        global b, g, r, xpos, ypos, clicked
        # clicked = True
        xpos = x
        ypos = y
        imgHSV = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        h, s, v = imgHSV[y, x]
        print(h,s,v)
        data_hsv = {'h':[h],
                    's':[s],
                    'v':[v]}
        
        df_hsv = pd.DataFrame(data_hsv)
        print(df_hsv)
        #df_hsv.to_excel('hsv_GD_C.xlsx', index=False) #create excel
        path = "hsv_GD_C.xlsx"
        book = load_workbook(path)
        writer = pd.ExcelWriter("hsv_GD_C.xlsx", engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df_hsv.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
        print("saving as hsv...")
        writer.save()
        

while(True):

    ret, frame = cap.read()

    frame40 = rescale_frame(frame, percent=40)
    cv2.imshow('frame40', frame40)
    cv2.setMouseCallback('frame40', draw_function, frame40)
    # This also acts as
    keyCode = cv2.waitKey(1) & 0xFF == ord('q')
    # Stop the program on the ESC key
    if keyCode == 27:
        break
    
    #cv2.imshow('image', imgscale)
    cv2.waitKey(1) 
    
cap.release()
cv2.destroyAllWindows()