from time import sleep
import numpy as np
import cv2
import math
import pyrealsense2 as rs
import pandas as pd
from openpyxl import load_workbook

# Prettier prints for reverse-engineering
from pprint import pprint


##################### PARAMETERS #########################
x,y = 0,0
points = [(0,0), (0,0)]
a,b,c,d,e,f,g,h,area1Ax,area1Ay,area1Bx,area1By,area2Ax,area2Ay,area2Bx,area2By,area3Ax,area3Ay,area3Bx,area3By,area4Ax,area4Ay,area4Bx,area4By,area5Ax,area5Ay,area5Bx,area5By= 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0
count = 0
##########################################################

#kp.init()
# Get realsense pipeline handle
pipe = rs.pipeline()

# Print all connected devices and find the T265
devices = rs.context().devices
for i in range(len(devices)):
    print('Found device:', devices[i].get_info(rs.camera_info.name), ', with serial number: ', devices[i].get_info(rs.camera_info.serial_number))

# Configure the pipeline
cfg = rs.config()

# Prints a list of available streams, not all are supported by each device
print('Available streams:')
pprint(dir(rs.stream))

# Enable streams you are interested in
cfg.enable_stream(rs.stream.pose) # Positional data (translation, rotation, velocity etc)
cfg.enable_stream(rs.stream.fisheye, 1) # Left camera
cfg.enable_stream(rs.stream.fisheye, 2) # Right camera

# Start the configured pipeline
pipe.start(cfg)

#################################### Realsense Input #################################################
def getRealsenseInput():
    global x,y
    frames = pipe.wait_for_frames()

    # Left fisheye camera frame
    left = frames.get_fisheye_frame(1)
    left_data = np.asanyarray(left.get_data())

    # Right fisheye camera frame
    right = frames.get_fisheye_frame(2)
    right_data = np.asanyarray(right.get_data())

    #print('Left frame', left_data.shape)
    #print('Right frame', right_data.shape)
    
    # Positional data frame
    pose = frames.get_pose_frame()
    if pose:
        pose_data = pose.get_pose_data()
        x = math.ceil(pose_data.translation.x * 100)
        y = math.ceil(pose_data.translation.z * 100)
        
        #print("Position xyz: % 2.4f % 2.4f" % (pose_data.translation.x, pose_data.translation.y))
        #print("Position x:  % 2.4f"%(x)) #2.4f artinya memberikan float, 4 angka dibelakang koma
        #print("Position y: % 2.4f"%(y))
    
    #x += int(fb)
    #y += int(lr)

    return [x, y]
########################################################################################################


###################################### Draw Pointer ####################################################
def drawPoints(img, points):
    #point = np.round(points).astype("int")
    for point in points:
        cv2.circle(img, point, 1, (150, 150, 150), cv2.FILLED) #tracking pointer
    cv2.circle(img, points[-1], 5, (0,0,255), cv2.FILLED)
    #cv2.circle(img, (x, y), 5, (0, 0, 255), cv2.FILLED) #pointer pada frame
    #cv2.circle(img, points, 5, (0,0,255), cv2.FILLED)
    #posX = (points[0])
    #posY = (points[1])
    cv2.putText(img, '(' + str(x) + ',' + str(y) + ')', ##penentuan titik #dikurangi 500 karena titik awal 500, dibagi 100 karena ingin mendapatkan dalam meter
        (points[-1][0]+10,points[-1][1]+30), cv2.FONT_HERSHEY_PLAIN, 1, #penempatan kalimatnya
        (255,0,255), 1)
    #print(posX, posY)
########################################################################################################


###################################### Draw Gedung ####################################################
def drawHome():
    cv2.rectangle(img, ((-50+500)/2, (-50+1000)/2), ((50+500)/2, (50+1000)/2), (255,255,255), cv2.FILLED) #+500 agar x di tengah +1000 agar z di tengah
def drawGdB():
    cv2.rectangle(img, ((-100+500)/2, (-900+1000)/2), ((100+500)/2, (-750+1000)/2), (0,255,0), cv2.FILLED)
def drawGdA():
    cv2.rectangle(img, ((-500+500)/2, (-900+1000)/2), ((-300+500)/2, (-700+1000)/2), (0,255,0), cv2.FILLED)
def drawGdC():
    cv2.rectangle(img, ((300+500)/2, (-900+1000)/2), ((500+500)/2, (-700+1000)/2), (0,255,0), cv2.FILLED)
def drawGD():
    drawHome()
    drawGdB()
    drawGdA()
    drawGdC()
########################################################################################################

def drawCoordinate(event, x, y, flags, frame):
    global a,b,c,d,e,f,g,h,area1Ax,area1Ay,area1Bx,area1By,area2Ax,area2Ay,area2Bx,area2By,area3Ax,area3Ay,area3Bx,area3By,area4Ax,area4Ay,area4Bx,area4By,area5Ax,area5Ay,area5Bx,area5By,count
    if event == cv2.EVENT_LBUTTONUP:
        nameGd = int(input('Where is the Buildings? '))
        if nameGd == 1:
            a = int(vals[0])
            b = int(vals[1])
        if nameGd == 2:
            c = int(vals[0])
            d = int(vals[1])
        if nameGd == 3:
            e = int(vals[0])
            f = int(vals[1])
        if nameGd == 4:
            g = int(vals[0])
            h = int(vals[1]) #gedung
        if nameGd == 5 :
            area1Ax = int(vals[0])
            area1Ay = int(vals[1]) #area
        if nameGd == 6 :
            area1Bx = int(vals[0])
            area1By = int(vals[1])
        if nameGd == 7 :
            area2Ax = int(vals[0])
            area2Ay = int(vals[1])
        if nameGd == 8:
            area2Bx = int(vals[0])
            area2By = int(vals[1])
        if nameGd == 9 :
            area3Ax = int(vals[0])
            area3Ay = int(vals[1])
        if nameGd == 10:
            area3Bx = int(vals[0])
            area3By = int(vals[1])
        if nameGd == 11 :
            area4Ax = int(vals[0])
            area4Ay = int(vals[1])
        if nameGd == 12 :
            area4Bx = int(vals[0])
            area4By = int(vals[1])
        if nameGd == 13 :
            area5Ax = int(vals[0])
            area5Ay = int(vals[1])
        if nameGd == 14 :
            area5Bx = int(vals[0])
            area5By = int(vals[1])
        data_Gd = {'Name':['Gd A','Gd B','Gd C','Obstacle','Area 1A','Area 1B', 'Area 2A', 'Area 2B','Area 3A', 'Area 3B','Area 4A','Area 4B','Area 5A','Area 5B'],
                    'CoordX':[a, c, e, g, area1Ax,area1Bx,area2Ax,area2Bx,area3Ax,area3Bx,area4Ax,area4Bx,area5Ax,area5Bx],
                    'CoordY':[b, d, f, h, area1Ay,area1By,area2Ay,area2By,area3Ay,area3By,area4Ay,area4By,area5Ay,area5By]}
        
        df_gd = pd.DataFrame(data_Gd) # tidak perlu kasih column name karena sudah ambil dari key yang ada di dalam dict
        if nameGd == 20 :
            #df_gd.to_excel('coordinate and area.xlsx', index=False)
            path = "coordinate and area.xlsx"
            book = load_workbook(path)
            writer = pd.ExcelWriter("coordinate and area.xlsx", engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df_gd.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
            print("saving as coordinate...")
            writer.save()
            count = 21
        print("Your Location = ", nameGd , int(vals[0]),int(vals[1]))
        print ("count: ", count)
        print(df_gd)
        
        

while True:
    #me.send_rc_control(vals[0], vals[1], vals[2], vals[3])
    img = np.zeros((550, 500, 3), np.uint8)
    drawGD()
    vals = getRealsenseInput()
    #print(vals[0], vals[1])
    
    #if (points[0] != vals[0] or points[1] != vals[1]):
     #get value from getKeyboardInput
    x = np.round(vals[0]).astype("int")
    y = np.round(vals[1]).astype("int")
    #print(x, y)
    points.append(((x+500)/2, (y+1000)/2))
    #points = ((vals[0]+500)/2, (vals[1]+1000)/2)
    #print(points)
    drawPoints(img, points)
    cv2.setMouseCallback('Output', drawCoordinate, img)
    cv2.imshow("Output", img)
    cv2.waitKey(1)